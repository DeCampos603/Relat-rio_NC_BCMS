#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório de Notas de Crédito — BCMS
==========================================================================
Relatório diário específico do Batalhão de Comunicações e Material Sigiloso,
contemplando as duas fontes executoras:

    • UASG 160329 — OGU (Orçamento Geral da União)
    • UASG 167329 — FEx (Fundo do Exército)

Módulos:
    1) Créditos RECEBIDOS  — dia anterior e semana corrente
    2) Créditos RECOLHIDOS — dia anterior e semana corrente
    3) Resumo consolidado  — totais e saldo líquido (OGU x FEx)

Fonte de dados: Google Sheets (export XLSX) alimentada do Tesouro Gerencial.
Entrega: Gmail SMTP (texto no formato WhatsApp com emojis).
Agendamento: GitHub Actions (dias úteis).
==========================================================================
"""

import os
import sys
import ssl
import smtplib
import traceback
import time
import re
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.request import urlretrieve
from urllib.error import HTTPError

import openpyxl

# =========================================================================
# CONFIGURAÇÃO
# =========================================================================

# ID hardcoded intencionalmente (lição aprendida na sprint anterior:
# secret SPREADSHEET_ID do GitHub foi historicamente mal configurado e
# causava HTTP 404 silencioso). A planilha é publicada por link — não
# há dado sensível em expor o ID neste arquivo.
SPREADSHEET_ID = "1Jv546wpWQSFAlep3oLRAg29hVy86iJxJ"
EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465

EMAIL_REMETENTE = os.environ.get("EMAIL_REMETENTE", "")
EMAIL_SENHA = os.environ.get("EMAIL_SENHA", "")
EMAIL_DESTINO = os.environ.get("EMAIL_DESTINO", "")
EMAIL_BCC = os.environ.get("EMAIL_BCC", "")

# -------------------------------------------------------------------------
# Mapeamento de colunas (Tesouro Gerencial — validado 23/04/2026)
# -------------------------------------------------------------------------
COL_UG_COD = 3      # UG Executora (UASG 6 dígitos)
COL_UG_NOME = 4     # UG Executora (nome)
COL_NC_ID = 5       # Número da NC
COL_ACAO = 6        # Ação Governo
COL_PI = 7          # Programa Interno (código)
COL_PI_DESC = 8     # PI (descrição)
COL_ND = 9          # Natureza Despesa (código)
COL_ND_DESC = 10    # ND (descrição)
COL_NC_DESC = 11    # NC - Descrição
COL_NC_TIPO = 12    # NC - Operação (Tipo)
COL_NC_DATA = 13    # NC - Dia Emissão
COL_CC = 15         # PROVISÃO RECEBIDA (valor CC em R$)
HEADER_ROWS = 8

# -------------------------------------------------------------------------
# Filtros do BCMS (este é o ponto que distingue do relatório multi-UG)
# -------------------------------------------------------------------------
UASGS_BCMS = {
    "160329": "OGU",
    "167329": "FEx",
}

FONTE_LABEL = {"160329": "OGU", "167329": "FEx"}
FONTE_ICON = {"160329": "\U0001F4BC", "167329": "\U0001F3E6"}  # 💼 🏦
FONTE_ORDEM = ["160329", "167329"]

# -------------------------------------------------------------------------
# Tipos de operação do SIAFI (campo "NC - Operação (Tipo)")
# -------------------------------------------------------------------------
TIPO_RECEBIDA = "DESCENTRALIZACAO DE CREDITO"
TIPOS_RECOLHIDA = {
    "ANULACAO DE DESCENTRALIZACAO DE CREDITO",
    "DEVOLUCAO DE DESCENTRALIZACAO DE CREDITO",
}
# DETALHAMENTO DE CREDITO é ignorado — é operação interna da própria UG
# sobre crédito já recebido, não é uma nova NC de entrada ou saída.

SALDO_MINIMO = 1.0  # ignora residuais < R$ 1 (artefatos SIGA de R$ 0,01)

# -------------------------------------------------------------------------
# Ícones do relatório
# -------------------------------------------------------------------------
ICON_REL = "\U0001F4C4"        # 📄
ICON_REC = "\U0001F4E5"        # 📥
ICON_DEV = "\U0001F4E4"        # 📤
ICON_DIA = "\U0001F4C5"        # 📅
ICON_SEM = "\U0001F4C6"        # 📆
ICON_TOT = "\U0001F4CA"        # 📊
ICON_OK_LINE = "\u2705"        # ✅
ICON_ERR_LINE = "\u274C"       # ❌
ICON_GREEN = "\U0001F7E2"      # 🟢
ICON_RED = "\U0001F534"        # 🔴
ICON_INFO = "\u2139\uFE0F"     # ℹ️
ICON_SEP = "\u2550"            # ═
ICON_DASH = "\u2500"           # ─


# =========================================================================
# UTILIDADES
# =========================================================================

def fmt_brl(v: float) -> str:
    """Formata valor em BRL: 1234.5 -> R$ 1.234,50 (com sinal se negativo)."""
    if v is None:
        v = 0.0
    sinal = "-" if v < 0 else ""
    v = abs(v)
    s = f"{v:,.2f}"
    # troca separadores para padrão BR
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{sinal}R$ {s}"


def parse_data_br(v):
    """Aceita datetime, date ou string 'DD/MM/YYYY'. Retorna date ou None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or "NAO" in s.upper() or s in ("-9", "'-9"):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def dia_util_anterior(hoje: date) -> date:
    """Retorna o último dia útil anterior a 'hoje' (seg→sex anterior)."""
    d = hoje - timedelta(days=1)
    while d.weekday() >= 5:   # 5=sáb, 6=dom
        d -= timedelta(days=1)
    return d


def janela_semana(hoje: date):
    """Retorna (segunda, sexta) da semana corrente."""
    segunda = hoje - timedelta(days=hoje.weekday())   # weekday: seg=0
    sexta = segunda + timedelta(days=4)
    return segunda, sexta


# =========================================================================
# DOWNLOAD E EXTRAÇÃO
# =========================================================================

def baixar_planilha() -> "openpyxl.worksheet.worksheet.Worksheet":
    destino = os.path.join(SCRIPT_DIR, "nc_bcms.xlsx")
    print(f"[{datetime.now():%H:%M}] Baixando planilha...")
    print(f"[{datetime.now():%H:%M}] URL: {EXPORT_URL}")

    ultimo_erro = None
    for tent in range(1, 4):
        try:
            urlretrieve(EXPORT_URL, destino)
            break
        except HTTPError as e:
            ultimo_erro = e
            if e.code == 404:
                raise RuntimeError(
                    f"Planilha não encontrada (HTTP 404). "
                    f"Verifique o SPREADSHEET_ID='{SPREADSHEET_ID}'."
                ) from e
            if tent < 3:
                espera = 2 ** tent   # 2s, 4s
                print(f"  Tent {tent} falhou ({e.code}); retry em {espera}s...")
                time.sleep(espera)
        except Exception as e:
            ultimo_erro = e
            if tent < 3:
                time.sleep(2 ** tent)
    else:
        raise RuntimeError(f"Download falhou após 3 tentativas: {ultimo_erro}")

    wb = openpyxl.load_workbook(destino, data_only=True)
    ws = wb.active
    print(f"[{datetime.now():%H:%M}] Planilha: {ws.max_row} linhas x {ws.max_column} colunas")
    return ws


def extrair_ncs_bcms(ws) -> list:
    """
    Extrai NCs das duas UASGs do BCMS (160329 e 167329).
    Filtra apenas linhas reais (exclui 'NAO SE APLICA' e detalhamentos).

    Retorna lista de dicts com:
      uasg, fonte ('OGU'|'FEx'), nc_id, acao, pi_cod, pi_desc,
      nd, nd_desc, descricao, tipo_op, data, cc, eh_recebida, eh_recolhida
    """
    registros = []

    for r in range(HEADER_ROWS + 1, ws.max_row + 1):
        uasg = str(ws.cell(r, COL_UG_COD).value or "").strip()
        if uasg not in UASGS_BCMS:
            continue

        nc_raw = str(ws.cell(r, COL_NC_ID).value or "").strip()
        if not nc_raw or nc_raw in ("'-9", "-9") or "NAO SE" in nc_raw.upper():
            continue

        tipo_op = str(ws.cell(r, COL_NC_TIPO).value or "").strip().upper()
        eh_recebida = (tipo_op == TIPO_RECEBIDA)
        eh_recolhida = (tipo_op in TIPOS_RECOLHIDA)
        if not eh_recebida and not eh_recolhida:
            continue   # ignora DETALHAMENTO e quaisquer outros

        data = parse_data_br(ws.cell(r, COL_NC_DATA).value)
        if data is None:
            continue

        cc_raw = ws.cell(r, COL_CC).value
        try:
            cc = float(cc_raw) if cc_raw not in (None, "", "-") else 0.0
        except (ValueError, TypeError):
            cc = 0.0
        if abs(cc) < SALDO_MINIMO:
            continue

        descricao = str(ws.cell(r, COL_NC_DESC).value or "").strip()
        acao = str(ws.cell(r, COL_ACAO).value or "").strip()
        pi_cod = str(ws.cell(r, COL_PI).value or "").strip()
        pi_desc = str(ws.cell(r, COL_PI_DESC).value or "").strip()
        nd = str(ws.cell(r, COL_ND).value or "").strip()
        nd_desc = str(ws.cell(r, COL_ND_DESC).value or "").strip()

        # Sanitiza campos “NAO SE APLICA” / '-9' do SIAFI
        if pi_cod in ("-9", "'-9"):
            pi_cod = ""
        if pi_desc in ("-9", "'-9") or "NAO SE" in pi_desc.upper():
            pi_desc = ""

        registros.append({
            "uasg": uasg,
            "fonte": FONTE_LABEL[uasg],
            "nc_id": nc_raw,
            "acao": acao,
            "pi_cod": pi_cod,
            "pi_desc": pi_desc,
            "nd": nd,
            "nd_desc": nd_desc,
            "descricao": descricao,
            "tipo_op": tipo_op,
            "data": data,
            "cc": cc,
            "eh_recebida": eh_recebida,
            "eh_recolhida": eh_recolhida,
        })

    print(f"[{datetime.now():%H:%M}] NCs BCMS extraídas: {len(registros)}")
    return registros


# =========================================================================
# RENDERIZAÇÃO
# =========================================================================

def _render_nc(nc: dict) -> list:
    """Renderiza uma NC em 2-3 linhas compactas (formato WhatsApp)."""
    linhas = []
    valor = fmt_brl(abs(nc["cc"]))
    nd_tag = f" _(ND {nc['nd']})_" if nc["nd"] else ""
    linhas.append(f"   • *{nc['nc_id']}*")
    linhas.append(f"      💰 {valor}{nd_tag}")
    if nc["pi_cod"]:
        pi_txt = f"PI {nc['pi_cod']}"
        if nc["pi_desc"]:
            pi_txt += f" — {nc['pi_desc'][:45]}"
        linhas.append(f"      📌 {pi_txt}")
    if nc["descricao"]:
        desc = nc["descricao"][:95]
        linhas.append(f"      📝 {desc}")
    return linhas


def _bloco_por_fonte(ncs: list, titulo_icon: str) -> list:
    """
    Monta dois blocos (OGU, FEx) para um conjunto de NCs.
    Se um bloco ficar vazio, emite "(sem NCs)" de forma discreta.
    """
    linhas = []
    for uasg in FONTE_ORDEM:
        fonte_ncs = [n for n in ncs if n["uasg"] == uasg]
        total_fonte = sum(abs(n["cc"]) for n in fonte_ncs)
        icon = FONTE_ICON[uasg]
        label = FONTE_LABEL[uasg]

        if not fonte_ncs:
            linhas.append(f"\n  {icon} *{label}* (UASG {uasg}) — _sem NCs no período_")
            continue

        linhas.append(
            f"\n  {icon} *{label}* (UASG {uasg}) — "
            f"{len(fonte_ncs)} NC{'s' if len(fonte_ncs) > 1 else ''} | "
            f"Total: *{fmt_brl(total_fonte)}*"
        )
        # ordena por data desc, depois por valor desc
        fonte_ncs_sorted = sorted(
            fonte_ncs, key=lambda n: (n["data"], abs(n["cc"])), reverse=True
        )
        for nc in fonte_ncs_sorted:
            linhas.extend(_render_nc(nc))
    return linhas


def _totais(ncs: list) -> dict:
    """Calcula totais separados por fonte."""
    out = {"total": 0.0, "OGU": 0.0, "FEx": 0.0}
    for n in ncs:
        v = abs(n["cc"])
        out["total"] += v
        out[n["fonte"]] += v
    return out


def gerar_relatorio(ncs: list, hoje: date, dia_ant: date,
                    segunda: date, sexta: date) -> str:
    """Gera o texto completo do relatório."""
    # Filtros temporais
    recebidas_dia = [n for n in ncs if n["eh_recebida"] and n["data"] == dia_ant]
    recebidas_sem = [n for n in ncs if n["eh_recebida"] and segunda <= n["data"] <= sexta]
    recolhidas_dia = [n for n in ncs if n["eh_recolhida"] and n["data"] == dia_ant]
    recolhidas_sem = [n for n in ncs if n["eh_recolhida"] and segunda <= n["data"] <= sexta]

    L = []
    sep = ICON_SEP * 42
    dash = ICON_DASH * 42

    # Cabeçalho
    L.append(f"{ICON_REL} *RELATÓRIO DE NOTAS DE CRÉDITO — BCMS*")
    L.append(f"📅 Referência: {hoje.strftime('%d/%m/%Y (%A)')}")
    L.append(f"💼 UASG OGU: 160329   |   🏦 UASG FEx: 167329")
    L.append("")

    # ─── MÓDULO 1 — RECEBIDOS ───────────────────────────────────────
    L.append(sep)
    L.append(f"{ICON_REC} *MÓDULO 1 — CRÉDITOS RECEBIDOS*")
    L.append(sep)

    L.append(f"\n{ICON_DIA} *DIA ANTERIOR — {dia_ant.strftime('%d/%m/%Y')}*")
    L.append(dash)
    if recebidas_dia:
        L.extend(_bloco_por_fonte(recebidas_dia, ICON_REC))
        t = _totais(recebidas_dia)
        L.append(f"\n  🔸 _Subtotal dia:_ *{fmt_brl(t['total'])}* "
                 f"(OGU {fmt_brl(t['OGU'])} | FEx {fmt_brl(t['FEx'])})")
    else:
        L.append("  _Nenhuma NC recebida._")

    L.append(f"\n{ICON_SEM} *SEMANA — {segunda.strftime('%d/%m')} a {sexta.strftime('%d/%m/%Y')}*")
    L.append(dash)
    if recebidas_sem:
        L.extend(_bloco_por_fonte(recebidas_sem, ICON_REC))
        t = _totais(recebidas_sem)
        L.append(f"\n  🔸 _Subtotal semana:_ *{fmt_brl(t['total'])}* "
                 f"(OGU {fmt_brl(t['OGU'])} | FEx {fmt_brl(t['FEx'])})")
    else:
        L.append("  _Nenhuma NC recebida na semana._")

    # ─── MÓDULO 2 — RECOLHIDOS ──────────────────────────────────────
    L.append("")
    L.append(sep)
    L.append(f"{ICON_DEV} *MÓDULO 2 — CRÉDITOS RECOLHIDOS / DEVOLVIDOS*")
    L.append(sep)

    L.append(f"\n{ICON_DIA} *DIA ANTERIOR — {dia_ant.strftime('%d/%m/%Y')}*")
    L.append(dash)
    if recolhidas_dia:
        L.extend(_bloco_por_fonte(recolhidas_dia, ICON_DEV))
        t = _totais(recolhidas_dia)
        L.append(f"\n  🔸 _Subtotal dia:_ *{fmt_brl(t['total'])}* "
                 f"(OGU {fmt_brl(t['OGU'])} | FEx {fmt_brl(t['FEx'])})")
    else:
        L.append("  _Nenhum recolhimento._")

    L.append(f"\n{ICON_SEM} *SEMANA — {segunda.strftime('%d/%m')} a {sexta.strftime('%d/%m/%Y')}*")
    L.append(dash)
    if recolhidas_sem:
        L.extend(_bloco_por_fonte(recolhidas_sem, ICON_DEV))
        t = _totais(recolhidas_sem)
        L.append(f"\n  🔸 _Subtotal semana:_ *{fmt_brl(t['total'])}* "
                 f"(OGU {fmt_brl(t['OGU'])} | FEx {fmt_brl(t['FEx'])})")
    else:
        L.append("  _Nenhum recolhimento na semana._")

    # ─── MÓDULO 3 — RESUMO CONSOLIDADO ──────────────────────────────
    L.append("")
    L.append(sep)
    L.append(f"{ICON_TOT} *MÓDULO 3 — RESUMO CONSOLIDADO*")
    L.append(sep)

    # Dia
    tr_dia = _totais(recebidas_dia)
    td_dia = _totais(recolhidas_dia)
    saldo_dia = tr_dia["total"] - td_dia["total"]
    saldo_dia_ogu = tr_dia["OGU"] - td_dia["OGU"]
    saldo_dia_fex = tr_dia["FEx"] - td_dia["FEx"]

    L.append(f"\n  {ICON_DIA} *Dia {dia_ant.strftime('%d/%m')}:*")
    L.append(f"    {ICON_OK_LINE} Recebido:  *{fmt_brl(tr_dia['total'])}*  "
             f"(OGU {fmt_brl(tr_dia['OGU'])} | FEx {fmt_brl(tr_dia['FEx'])})")
    L.append(f"    {ICON_ERR_LINE} Recolhido: *{fmt_brl(td_dia['total'])}*  "
             f"(OGU {fmt_brl(td_dia['OGU'])} | FEx {fmt_brl(td_dia['FEx'])})")
    icon_s = ICON_GREEN if saldo_dia >= 0 else ICON_RED
    L.append(f"    {icon_s} Saldo:     *{fmt_brl(saldo_dia)}*  "
             f"(OGU {fmt_brl(saldo_dia_ogu)} | FEx {fmt_brl(saldo_dia_fex)})")

    # Semana
    tr_sem = _totais(recebidas_sem)
    td_sem = _totais(recolhidas_sem)
    saldo_sem = tr_sem["total"] - td_sem["total"]
    saldo_sem_ogu = tr_sem["OGU"] - td_sem["OGU"]
    saldo_sem_fex = tr_sem["FEx"] - td_sem["FEx"]

    L.append(f"\n  {ICON_SEM} *Semana {segunda.strftime('%d/%m')}–{sexta.strftime('%d/%m')}:*")
    L.append(f"    {ICON_OK_LINE} Recebido:  *{fmt_brl(tr_sem['total'])}*  "
             f"(OGU {fmt_brl(tr_sem['OGU'])} | FEx {fmt_brl(tr_sem['FEx'])})")
    L.append(f"    {ICON_ERR_LINE} Recolhido: *{fmt_brl(td_sem['total'])}*  "
             f"(OGU {fmt_brl(td_sem['OGU'])} | FEx {fmt_brl(td_sem['FEx'])})")
    icon_s = ICON_GREEN if saldo_sem >= 0 else ICON_RED
    L.append(f"    {icon_s} Saldo:     *{fmt_brl(saldo_sem)}*  "
             f"(OGU {fmt_brl(saldo_sem_ogu)} | FEx {fmt_brl(saldo_sem_fex)})")

    # Rodapé
    L.append("")
    L.append(dash)
    L.append(f"{ICON_INFO} _Fonte: Tesouro Gerencial → Google Sheets → SIAFI._")
    L.append(f"_Gerado automaticamente em {datetime.now():%d/%m/%Y %H:%M} "
             f"via GitHub Actions._")

    return "\n".join(L)


# =========================================================================
# ENVIO DE E-MAIL
# =========================================================================

def _parse_lista(s: str) -> list:
    """Separa por ',' ou ';' e remove espaços/vazios."""
    if not s:
        return []
    partes = re.split(r"[,;]", s)
    return [p.strip() for p in partes if p.strip()]


def enviar_email(corpo: str, assunto: str) -> None:
    if not (EMAIL_REMETENTE and EMAIL_SENHA and EMAIL_DESTINO):
        print(f"[{datetime.now():%H:%M}] Credenciais ausentes — imprimindo localmente:\n")
        print(corpo)
        return

    dest_to = _parse_lista(EMAIL_DESTINO)
    dest_bcc = _parse_lista(EMAIL_BCC)
    envelope = dest_to + dest_bcc   # entrega real (SMTP RCPT TO)

    if not envelope:
        raise RuntimeError("EMAIL_DESTINO vazio após parse.")

    msg = MIMEMultipart("alternative")
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = ", ".join(dest_to)
    msg["Subject"] = assunto
    # NÃO incluir cabeçalho Bcc: no MIME — o Bcc entra só no envelope.

    msg.attach(MIMEText(corpo, "plain", "utf-8"))

    context = ssl.create_default_context()
    print(f"[{datetime.now():%H:%M}] Conectando {SMTP_SERVER}:{SMTP_PORT}...")
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as smtp:
        smtp.login(EMAIL_REMETENTE, EMAIL_SENHA)
        smtp.sendmail(EMAIL_REMETENTE, envelope, msg.as_string())

    print(f"[{datetime.now():%H:%M}] E-mail enviado.")
    print(f"  To:  {dest_to}")
    if dest_bcc:
        print(f"  Bcc: {dest_bcc}")


# =========================================================================
# MAIN
# =========================================================================

def main() -> None:
    try:
        hoje = date.today()
        dia_ant = dia_util_anterior(hoje)
        segunda, sexta = janela_semana(hoje)

        print(f"\n{'='*60}")
        print(f"  RELATÓRIO DE NCs — BCMS")
        print(f"  Execução: {datetime.now():%d/%m/%Y %H:%M:%S}")
        print(f"  Dia anterior (útil): {dia_ant}")
        print(f"  Semana: {segunda} a {sexta}")
        print(f"{'='*60}\n")

        ws = baixar_planilha()
        ncs = extrair_ncs_bcms(ws)

        corpo = gerar_relatorio(ncs, hoje, dia_ant, segunda, sexta)
        assunto = f"[BCMS] Relatório de NCs — {hoje.strftime('%d/%m/%Y')}"

        enviar_email(corpo, assunto)
        print(f"\n[{datetime.now():%H:%M}] Concluído com sucesso.")

    except Exception:
        print(f"\n[{datetime.now():%H:%M}] ERRO durante a execução:")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
